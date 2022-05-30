from threading import Thread
from time import time

from sqlalchemy import create_engine
from sqlalchemy import text
import re
import psycopg2
import openpyxl

engine_from = create_engine("postgresql+psycopg2://dedul:dedul@localhost:15432/gis")


def __sort_results(results):
    return sorted(results, key=lambda row: (
        int(re.search(r'\d+', row['housenumb']).group()) if row['housenumb'] is not None and re.search(r'\d+', row[
                                                                'housenumb']) is not None else 0
    ), reverse=True)


def __search_address(error_address_before, error_address_after, result, housenumber):
    if not housenumber:
        return result
    list_addr = []
    for index, dict_ in enumerate(result):
        try:
            s = __get_int_type(dict_['housenumb']) - __get_int_type(housenumber)
            if error_address_before <= s <= error_address_after:
                list_addr.append(dict_)
        except:
            pass
    return list_addr


def __search_address2(error_address_before, error_address_after, result, housenumber):
    return list(filter(lambda item: error_address_before < __get_int_type(item['housenumb']) - __get_int_type(
        housenumber) < error_address_after, result))


def __get_int_type(house_number):
    return int(re.search(r'\d+', house_number).group()) if house_number is not None \
                                                           and re.search(r'\d+',
                                                                         house_number) is not None else 0


def correctly_address(results, housenumber):
    full_compare_numb_address = None
    compare_numb_address = None
    for index, dict_ in enumerate(results):
        numb_out = __get_int_type(dict_['housenumb'])
        numb_input = __get_int_type(housenumber)

        if str(dict_['housenumb']) == str(housenumber.strip()):
            full_compare_numb_address = dict_
        elif numb_out == numb_input:
            if full_compare_numb_address is None:
                compare_numb_address = dict_

    if full_compare_numb_address is not None:
        return full_compare_numb_address
    elif compare_numb_address is not None:
        return compare_numb_address
    else:
        return None


def correctly_address2(results, housenumber):
    full_compare_numb_address = next((x for x in results if str(x['housenumb']) == str(housenumber)), None)

    if full_compare_numb_address is not None:
        return full_compare_numb_address
    else:
        return next((x for x in results if __get_int_type(x['housenumb']) == __get_int_type(housenumber)), None)


def not_correctly_address(results, housenumber, street):
    if not results:
        return {}

    if street is None:
        return results[0]

    if housenumber is None or housenumber == "":
        return results[0]

    left = __get_int_type(housenumber)
    right = __get_int_type(results[0]['housenumb'])

    n = left - right
    if n < 0:
        n, m = left - right, right - left
    else:
        m, n = left - right, right - left
    result_arr = results
    while len(result_arr) > 4:
        result_arr = __search_address(n, m, result_arr, housenumber)
        n += 1
        m -= 1

    if result_arr:
        return result_arr[0]

    return results[0]


def get_correct_housenumber(number):
    housenumber = ''
    if number != '' and number is not None:
        number = number.strip()
        if number.find('-') > 0:
            for n in number.split('-'):
                tmp_n = re.search(r'\d+', n).group() if n is not None and re.search(r'\d+', n) is not None else ""
                if tmp_n.isnumeric():
                    housenumber = n
                    break
        else:
            housenumber = number
        return housenumber.strip()
    else:
        return None


def __check_symbols_drop_in_word(addr):
    if addr.find(".") > 0:
        s = re.findall(r'\.\w+$', addr)
        return s[0].replace(".", '') if s else addr.replace(".", '')
    return addr


def __check_symbols_dash_in_word(addr):
    if addr.find('-') == 1:
        return re.findall(r'^\w+[\-]', addr)[0].replace("-", '')
    else:
        return __check_symbols_drop_in_word(addr.replace("-", ''))


def get_correct_address(address):
    if address == "Неизвестная":
        address = None
    if isinstance(address, str) and address != "":
        address = address.strip().replace(",", " ")
        correct_addr_list = []
        if address.find(' ') > 0:
            for addr in address.split(' '):
                s = __check_symbols_drop_in_word(__check_symbols_dash_in_word(addr))
                correct_addr_list.append(s)
            return correct_addr_list
        else:
            s = __check_symbols_drop_in_word(__check_symbols_dash_in_word(address))
            return s
    else:
        return None


def __create_sql_query_for_city(city, template=None):
    city_query = ""
    if city:
        if isinstance(city, str):
            # city_query = f"WHERE name = '{city}'"
            city_query = template
        elif isinstance(city, list):
            city_query = "WHERE translate(name, 'ё', 'е') like translate('%" + city[0] \
                         + "%', 'ё', 'е') "
            for i in range(1, len(city)):
                city_query += "and translate(name, 'ё', 'е') like translate('%" + city[i] \
                              + "%', 'ё', 'е')  "
        return city_query
    else:
        return None


def __create_sql_query_for_street(street):
    street_query = ""
    if street:
        if isinstance(street, str):
            street_query = f"WHERE translate(lower(tags->'addr:street'), 'ёо', 'еа') " \
                           f"like translate(lower('%{street}%'), 'ёо', 'еа') "
        elif isinstance(street, list):
            if street:
                street_query = f"WHERE translate(lower(tags->'addr:street'), 'ёо', 'еа') " \
                               f"like translate(lower('%{street[0]}%'), 'ёо', 'еа') "
                for i in range(1, len(street)):
                    street_query += f"and translate(lower(tags->'addr:street'), 'ёо', 'еа') " \
                                    f"like translate(lower('%{street[i]}%'), 'ёо', 'еа') "
        return street_query
    else:
        return ""


def __get_addresses_list_from_db(city, street=None, template=None):
    results = []
    city_query = __create_sql_query_for_city(city, template)
    street_query = __create_sql_query_for_street(street)
    if city_query:
        sql = text("select lat, long, housenumber, housename, regionname from ( \
                            select way_o, way_ul, lat, long, housenumber, housename, tags, regionname, ST_Contains(way_o,way_ul) ch  from ( \
                            SELECT way way_o, name as regionname\
                            FROM planet_osm_polygon \
                            " + city_query + " \
                            ) t1, \
                            (SELECT tags, \"addr:housenumber\" housenumber, way way_ul, \
                            replace(ST_X(ST_TRANSFORM(ST_Centroid(way),4674))::varchar(255),',','.') AS long, \
                            replace(ST_Y(ST_TRANSFORM(ST_Centroid(way), 4674))::varchar(255), ',', '.') AS lat, tags->'addr:street' as housename \
                            FROM planet_osm_polygon \
                            " + street_query + " \
                            ) t2 \
                            ) t5 \
                            where ch is True")
        result = engine_from.execute(sql)
        for row in result:
            results.append(dict(zip(['lat', 'lon', 'housenumb', 'housename', 'regionname'], row)))

        return results
    else:
        return []


def get_address_from_db(city, street, housenumber):
    addr = {}
    street_correct = get_correct_address(street)
    city_correct = get_correct_address(city)
    housenumber = get_correct_housenumber(housenumber)
    print(f"{city_correct} {street_correct} {housenumber}")
    results = __get_addresses_list_from_db(city_correct, street_correct, f"WHERE name = '{city_correct}'")
    if not results:
        results = __get_addresses_list_from_db(city_correct, street_correct, f"WHERE translate(name, 'ё', 'е') = "
                                                                             f"translate('{city_correct}', 'ё', "
                                                                             f"'е') ")
    if not results:
        results = __get_addresses_list_from_db(city_correct, street=None, template=f"WHERE name = '{city_correct}'")
        street_correct = None

    if not results:
        results = __get_addresses_list_from_db(city_correct, street=None, template=f"WHERE translate(name, 'ё', 'е') = "
                                                                                   f"translate('{city_correct}', 'ё', "
                                                                                   f"'е') ")

        street_correct = None
    if street_correct is not None:
        results = __sort_results(results)

    correctly = True

    if housenumber is not None and street_correct is not None:
        addr = correctly_address(results, housenumber)

    if not addr:
        addr = not_correctly_address(results, housenumber, street_correct)
        correctly = False

    if addr:
        return {
            'current_address': {
                'current_city': city,
                'current_street': street,
                'current_housenumber': housenumber
            },
            'found_address': {
                'lat': addr['lat'],
                'lan': addr['lon'],
                'found_city': addr['regionname'],
                'found_street': addr['housename'],
                'found_housenumber': addr['housenumb']
            },
            'correctly': correctly
        }
    else:
        return {
            'current_address': {
                'current_city': city,
                'current_street': street,
                'current_housenumber': housenumber},
            'found_address': {
                'address': 'not found'
            }
        }


def work_with_files(path):
    count = 0
    countNFList = []
    countFalse = 0
    countTrue = 0

    print(get_address_from_db('Жары', 'Центральная', '14'))

    return
    column = path['number'] + 9
    workbook = openpyxl.load_workbook('input/' + path['url'])
    worksheet = workbook.active

    # for i in range(1, 9):
    for i in range(1, worksheet.max_row + 1):
        # try:
        city = worksheet.cell(i, path['city']).value
        street = worksheet.cell(i, path['street']).value
        housenumber = worksheet.cell(i, path['number']).value
        print(f"{city} {street} {housenumber}")
        result = get_address_from_db(city, street, housenumber)
        print(result)
        try:
            worksheet.cell(row=i,
                           column=column).value = f"({result['found_address']['lat']}, {result['found_address']['lan']})"
            worksheet.cell(row=i, column=column + 1).value = f"({result['correctly']})"
        except KeyError:
            pass

        try:
            if not result['correctly']:
                countFalse += 1
            else:
                countTrue += 1
        except KeyError:
            pass

        try:
            if result['found_address']['address']:
                countNFList.append(result)
                count += 1
        except KeyError:
            pass
        # except:
        #     continue

    workbook.save(filename="output/" + path['url'])

    print(f"Not Found: {count}")
    print(f"False: {countFalse}")
    print(f"True: {countTrue}")

    f = open('out2.txt', 'w')
    for l in countNFList:
        f.write(f"print(get_address_from_db('{l['current_address']['current_city']}', "
                f"'{l['current_address']['current_street']}', "
                f"'{l['current_address']['current_housenumber']}'))\n")


def main():
    filenames = [
        # {
        #     'url': "Белкоопсоюз-Торговый-реестр.xlsx",
        #     'city': 7,
        #     'street': 8,
        #     'number': 9
        # },
        # {
        #     'url': "Белкоопсоюз-Торговый-реестр-Общепит.xlsx",
        #     'city': 7,
        #     'street': 8,
        #     'number': 9
        # },
        # {
        #     'url': "Белпочта-Торговый-реестр.xlsx",
        #     'city': 7,
        #     'street': 8,
        #     'number': 9
        # },
        # {
        #     'url': "Белсоюзпечать-Торговый-реестр.xlsx",
        #     'city': 7,
        #     'street': 8,
        #     'number': 9
        # },
        {
            'url': "Santa_read.xlsx",
            'city': 5,
            'street': 6,
            'number': 7
        },
        # {
        #     'url': "ЭНЕРГО-ОИЛ-Торговый-реестр-новые.xlsx",
        #     'city': 5,
        #     'street': 6,
        #     'number': 7
        # },
        # {
        #     'url': "Сведения_Торгового_реестра_Республики_Беларусь_на_25042022_текущий_2022_загрузка.xlsx",
        #     'city': 22,
        #     'street': 23,
        #     'number': 24
        # }
    ]

    # for i in filenames:
    #     th = Thread(target=work_with_files, args=(i,))
    #     th.start()
    # work_with_files(i)
    t0 = time()
    work_with_files(filenames[0])
    tend = (time() - t0)
    print("time : " + str(tend))


if __name__ == '__main__':
    main()
